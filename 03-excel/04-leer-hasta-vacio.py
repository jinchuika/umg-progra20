import openpyxl

book = openpyxl.load_workbook('excel_viejo.xls', data_only=True)

hoja = book.active

for fila in hoja.iter_rows(min_col=1, max_col=hoja.max_column):
    print([celda.value for celda in fila])
